# output_builder.py
# Builds the output Excel workbook with one sheet per DB table.
# Each sheet has EXACTLY the columns of the corresponding DB table in exact order.
# This workbook is the migration review artifact - check it before any real DB insert.
#
# Sheet order matches FK insert order:
# 1. center, 2. user, 3. patient, 4. visit,
# 5. prescription, 6. prescriptionitem, 7. followup_schedule
# + 1 extra sheet: migration_log (all cleaning issues found)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# - EXACT COLUMN DEFINITIONS - match new DB schema exactly --------------------
TABLE_COLUMNS = {
    "center": [
        "id", "name", "district", "state", "is_active", "center_code",
        "pincode", "phone", "email", "equipment", "created_at", "tenant_id",
        "short_name"
    ],
    "user": [
        "id", "role", "first_name", "last_name", "employee_id", "phone", "email",
        "password_hash", "center_id", "is_active", "created_at", "specialization",
        "qualifications", "is_available_now", "notes", "avatar_url", "is_away",
        "away_since", "fcm_token", "google_meet_link", "zoom_link",
        "preferred_video_platform", "signature_url", "licence_number", "pin_code",
        "per_consultation_fee", "other_charges", "government_id",
        "doctor_review_preference", "allowed_video_platforms", "tenant_id",
        "ltc_suggest_unplanned_chronic", "ltc_suggest_frequent_flyer",
        "legacy_id", "legacy_source"
    ],
    "patient": [
        "id", "full_name", "phone", "sex", "date_of_birth", "age_years",
        "address_text", "district", "state", "occupation", "category",
        "mother_tongue", "government_id", "center_id", "created_at",
        "address_line_1", "address_line_2", "village_town_city", "postal_code",
        "country", "address_resolution_method", "pincode_master_id", "age_months",
        "health_history_json", "long_term_conditions", "consent_status", "family_id",
        "patient_status", "is_demo", "deleted_at", "tenant_id", "phone_last10",
        "merged_into_patient_id", "legacy_id", "legacy_source",
        "legacy_extra_1", "legacy_extra_2", "legacy_extra_3", "migrated_at"
    ],
    "visit": [
        "id", "patient_id", "center_id", "created_by_user_id",
        "assigned_doctor_user_id", "status", "chief_complaint", "complaint_duration",
        "history_notes", "is_reconsultation", "vitals_json", "created_at",
        "updated_at", "completed_at", "camp_id", "resolved_followup_visit_id",
        "family_living_count_at_visit", "family_living_status_confirmed",
        "family_living_status_updated_during_visit", "triage_session_id", "outcome",
        "deleted_at", "tenant_id", "consultation_type", "legacy_id", "legacy_source",
        "resolved_followup_schedule_id"
    ],
    "prescription": [
        "id", "visit_id", "doctor_user_id", "assessment", "provisional_diagnosis",
        "confirmed_diagnosis", "is_referral_recommended", "referral_treatment_name",
        "referral_specialist_name", "instructions", "follow_up_at", "created_at",
        "investigation_notes", "deleted_at", "patient_instructions",
        "legacy_id", "legacy_source"
    ],
    "prescriptionitem": [
        "id", "prescription_id", "medication_name", "dosage", "frequency",
        "duration", "quantity", "unit", "notes", "formulary_id"
    ],
    "followup_schedule": [
        "id", "treatment_plan_id", "sequence_no", "scheduled_date", "status",
        "completed_visit_id", "assigned_coordinator_id", "reminder_sent_at",
        "rescheduled_from_id", "cancelled_reason", "notes", "created_at",
        "updated_at", "patient_id", "source_prescription_id", "resolution",
        "resolution_notes", "resolved_by_user_id", "resolved_at"
    ],
}

FK_COLUMNS = {
    "user": ["center_id"],
    "patient": ["center_id"],
    "visit": ["patient_id", "center_id", "created_by_user_id", "assigned_doctor_user_id"],
    "prescription": ["visit_id", "doctor_user_id"],
    "prescriptionitem": ["prescription_id"],
    "followup_schedule": ["patient_id", "source_prescription_id"],
}

NOT_NULL_COLUMNS = {
    "center": ["id", "name", "is_active"],
    "user": ["id", "role", "first_name", "last_name", "phone", "email", "password_hash", "is_active", "is_available_now", "is_away", "doctor_review_preference", "ltc_suggest_unplanned_chronic", "ltc_suggest_frequent_flyer"],
    "patient": ["id", "full_name", "created_at", "patient_status", "is_demo"],
    "visit": ["id", "patient_id", "center_id", "created_by_user_id", "status", "chief_complaint", "is_reconsultation", "vitals_json", "created_at", "updated_at"],
    "prescription": ["id", "visit_id", "doctor_user_id", "assessment", "is_referral_recommended", "created_at"],
    "prescriptionitem": ["id", "prescription_id", "medication_name", "dosage", "frequency"],
    "followup_schedule": ["id", "sequence_no", "scheduled_date", "status", "created_at", "updated_at", "patient_id"],
}


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FK_FILL = PatternFill("solid", fgColor="FDE68A")
NULL_FILL = PatternFill("solid", fgColor="FCA5A5")
ALT_FILL = PatternFill("solid", fgColor="F9FAFB")
LOG_HEADER_FILL = PatternFill("solid", fgColor="991B1B")


def dict_to_rows(data: list[dict], columns: list[str]) -> pd.DataFrame:
    """
    Convert list of row dicts to a DataFrame with exact column order.
    """
    if not data:
        return pd.DataFrame(columns=columns)
    frame = pd.DataFrame(data)
    for column in columns:
        if column not in frame.columns:
            frame[column] = None
    frame = frame.reindex(columns=columns)
    return frame


def _style_header(cell, is_fk: bool = False, is_log: bool = False):
    cell.fill = LOG_HEADER_FILL if is_log else (FK_FILL if is_fk else HEADER_FILL)
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    for column_cells in ws.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        max_len = max((len(value) for value in values), default=0)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_len + 2, 45)


def _append_summary_row(ws, total_rows: int, column_count: int):
    summary = [None] * column_count
    summary[0] = f"Total rows: {total_rows}"
    ws.append(summary)


def build_output_workbook(
    center_rows: list[dict],
    user_rows: list[dict],
    patient_rows: list[dict],
    visit_rows: list[dict],
    prescription_rows: list[dict],
    prescriptionitem_rows: list[dict],
    followup_rows: list[dict],
    migration_log: list[dict],
    output_path: str,
):
    """
    Build the output Excel workbook.
    """
    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = "_temp"

    table_data = {
        "center": center_rows,
        "user": user_rows,
        "patient": patient_rows,
        "visit": visit_rows,
        "prescription": prescription_rows,
        "prescriptionitem": prescriptionitem_rows,
        "followup_schedule": followup_rows,
    }

    for table_name, columns in TABLE_COLUMNS.items():
        ws = wb.create_sheet(table_name)
        frame = dict_to_rows(table_data.get(table_name, []), columns)

        for row_index, row in enumerate(dataframe_to_rows(frame, index=False, header=True), start=1):
            ws.append(list(row))
            if row_index == 1:
                for col_index, column_name in enumerate(columns, start=1):
                    header_cell = ws.cell(1, col_index)
                    _style_header(header_cell, is_fk=column_name in FK_COLUMNS.get(table_name, []))
                    if column_name in NOT_NULL_COLUMNS.get(table_name, []):
                        header_cell.value = f"{column_name}*"
            else:
                fill = ALT_FILL if row_index % 2 == 0 else None
                for col_index, column_name in enumerate(columns, start=1):
                    cell = ws.cell(row_index, col_index)
                    if fill is not None:
                        cell.fill = fill
                    if column_name in NOT_NULL_COLUMNS.get(table_name, []) and cell.value in (None, ""):
                        cell.fill = NULL_FILL

        _append_summary_row(ws, len(frame), len(columns))
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True
        _auto_width(ws)

    log_ws = wb.create_sheet("migration_log")
    log_columns = ["uid", "field", "issue", "raw", "detail"]
    log_frame = pd.DataFrame(migration_log)
    for column in log_columns:
        if column not in log_frame.columns:
            log_frame[column] = None
    log_frame = log_frame.reindex(columns=log_columns)
    if not log_frame.empty and "issue" in log_frame.columns:
        severity_order = {"NOT_NULL_VIOLATION": 0, "FK_MISSING": 1, "DUPLICATE_UUID": 2}
        log_frame["_severity"] = log_frame["issue"].map(lambda value: severity_order.get(value, 99))
        log_frame = log_frame.sort_values(by=["_severity", "uid", "field"], kind="stable").drop(columns=["_severity"])

    for col_index, column_name in enumerate(log_columns, start=1):
        cell = log_ws.cell(1, col_index, column_name)
        _style_header(cell, is_log=True)
    for row in log_frame.itertuples(index=False):
        log_ws.append(list(row))
    log_ws.freeze_panes = "A2"
    _auto_width(log_ws)

    wb.remove(default_sheet)
    wb.save(output_path)
